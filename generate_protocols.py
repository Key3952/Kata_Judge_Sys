import csv
import io
import json
import os
import shutil
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook

from csv_manager import CompetitionCSVManager, normalize_protocol_token, safe_float, safe_int

# Импортируем функцию для красивых названий дисциплин
def get_discipline_display_name(key: str) -> str:
    """Получает красивое название дисциплины"""
    display_names = {
        'nagenokata': 'Nage-no-kata',
        'katamenokata': 'Katame-no-kata',
        'kimenokata': 'Kime-no-kata',
        'junokata': 'Ju-no-kata',
        'kodokangoshinjutsu': 'Kodokan Goshin-jutsu',
        'koshikinokata': 'Koshiki-no-kata',
        'itsutsunokata': 'Itsutsu-no-kata',
    }
    return display_names.get(key.lower(), key)


def _participant_detail_line(pair_row: dict, prefix: str) -> str:
    """prefix: 'Тори_' или 'Уке_'"""
    parts = [
        pair_row.get(f'{prefix}год рождения', '').strip(),
        pair_row.get(f'{prefix}разряд', '').strip(),
        pair_row.get(f'{prefix}кю', '').strip(),
        pair_row.get(f'{prefix}СШ', '').strip(),
        pair_row.get(f'{prefix}тренер', '').strip(),
    ]
    return ', '.join(p for p in parts if p)


def encode_participant_for_protocol(pair_row: dict, role: str) -> str:
    """role: 'Тори' или 'Уке'. В CSV: Имя||остальное через запятую"""
    prefix = f'{role}_'
    name = pair_row.get(f'{prefix}ФИО', '').strip()
    detail = _participant_detail_line(pair_row, prefix)
    if detail:
        return f'{name}||{detail}'
    return name
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

SKIP_DIRS = frozenset({"__pycache__", "results"})

_PDF_FONT: Optional[str] = None
_PDF_FONT_BOLD: Optional[str] = None


def _ensure_pdf_fonts() -> Tuple[str, str]:
    global _PDF_FONT, _PDF_FONT_BOLD
    if _PDF_FONT is not None:
        return _PDF_FONT, _PDF_FONT_BOLD or "Helvetica-Bold"
    candidates = [
        "/usr/share/fonts/TTF/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ]
    for p in candidates:
        if os.path.isfile(p):
            bp = p.replace("DejaVuSans.ttf", "DejaVuSans-Bold.ttf")
            try:
                pdfmetrics.registerFont(TTFont("KataDejaVu", p))
                _PDF_FONT = "KataDejaVu"
                if os.path.isfile(bp):
                    pdfmetrics.registerFont(TTFont("KataDejaVu-Bold", bp))
                    _PDF_FONT_BOLD = "KataDejaVu-Bold"
                else:
                    _PDF_FONT_BOLD = "KataDejaVu"
                return _PDF_FONT, _PDF_FONT_BOLD
            except Exception:
                pass
    _PDF_FONT, _PDF_FONT_BOLD = "Helvetica", "Helvetica-Bold"
    return _PDF_FONT, _PDF_FONT_BOLD


def _discipline_folders(comp_path: str) -> List[str]:
    out = []
    if not os.path.isdir(comp_path):
        return out
    for name in sorted(os.listdir(comp_path)):
        p = os.path.join(comp_path, name)
        if os.path.isdir(p) and name not in SKIP_DIRS and name != "__pycache__":
            out.append(name)
    return out


def _read_csv_rows(path: str) -> List[Dict[str, str]]:
    if not os.path.exists(path):
        return []
    rows: List[Dict[str, str]] = []
    with open(path, "r", encoding="utf-8", errors="replace", newline="") as f:
        r = csv.DictReader(f)
        if r.fieldnames is None:
            return []
        for row in r:
            rows.append({k: (v or "") for k, v in row.items() if k is not None})
    return rows


def _load_stage(comp_path: str, discipline_key: str) -> Dict[str, str]:
    p = os.path.join(comp_path, discipline_key, "stage.json")
    cfg = {"current_stage": "final", "mode": "final_only", "status": "open"}
    if os.path.exists(p):
        try:
            with open(p, "r", encoding="utf-8") as f:
                cfg.update(json.load(f) or {})
        except Exception:
            pass
    return cfg


def _stage_paths(comp_path: str, discipline_key: str, stage: str) -> Dict[str, str]:
    s = "final" if str(stage).lower() == "final" else "prelim"
    disc_root = os.path.join(comp_path, discipline_key)
    base = os.path.join(disc_root, "final") if s == "final" else disc_root
    return {
        "pairs": os.path.join(base, "participants_list.csv"),
        "final": os.path.join(base, "final_protocol.csv"),
        "protocols": os.path.join(base, "protocols"),
    }


def _decode_participant_cell(cell_value: Any) -> Dict[str, str]:
    if cell_value is None:
        return {"name": "", "detail": ""}
    s = str(cell_value).strip()
    if "||" in s:
        a, _, b = s.partition("||")
        return {"name": a.strip(), "detail": b.strip()}
    return {"name": s, "detail": ""}


def _score_from_detail(d: dict) -> float:
    if not d:
        return 10.0
    if d.get("forgotten"):
        return 0.0
    v = 10.0
    v -= float(d.get("m1", 0) or 0)
    v -= float(d.get("m2", 0) or 0)
    v -= float(d.get("med", 0) or 0)
    v -= float(d.get("big", 0) or 0)
    v -= float(d.get("c_minus", 0) or 0)
    v -= float(d.get("c_plus", 0) or 0)
    return max(0.0, min(10.0, v))


def _judge_total(details: List[dict]) -> float:
    total = sum(_score_from_detail(d) for d in details)
    if any(d.get("forgotten") for d in details):
        total /= 2.0
    return total


def _safe_sheet_name(name: str, used: set) -> str:
    """Генерирует безопасное имя листа Excel (макс 31 символ, без спецсимволов)"""
    bad = set('[]:*?/\\')
    n = "".join("_" if ch in bad else ch for ch in name).strip() or "Sheet"
    n = n[:31]
    base = n
    i = 1
    while n in used:
        suffix = f"_{i}"
        n = (base[: 31 - len(suffix)] + suffix)
        i += 1
    used.add(n)
    return n


def _theme_gradient(score: float) -> str:
    """Генерирует цвет ячейки на основе оценки (градиент от красного к зеленому)"""
    # палитра как в табло (стандарт)
    p = [
        (230, 124, 115),
        (243, 169, 109),
        (255, 214, 102),
        (171, 201, 120),
        (87, 187, 138),
    ]
    t = max(0.0, min(1.0, safe_float(score, 0.0) / 170.0))
    seg = len(p) - 1
    idx = min(seg - 1, int(t * seg))
    lt = t * seg - idx
    r = round(p[idx][0] + (p[idx + 1][0] - p[idx][0]) * lt)
    g = round(p[idx][1] + (p[idx + 1][1] - p[idx][1]) * lt)
    b = round(p[idx][2] + (p[idx + 1][2] - p[idx][2]) * lt)
    return f"{r:02X}{g:02X}{b:02X}"


def _autosize(ws):
    for col in ws.columns:
        m = 0
        l = get_column_letter(col[0].column)
        for c in col:
            if c.value is not None:
                m = max(m, len(str(c.value)))
        ws.column_dimensions[l].width = min(max(10, m + 2), 60)


def _apply_grid(ws, start_row: int, end_row: int, start_col: int, end_col: int):
    thin = Side(style="thin", color="6B7280")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(r, c)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _pair_slug_from_participants_row(p: dict) -> str:
    t = normalize_protocol_token(p.get("Тори_ФИО", ""))
    u = normalize_protocol_token(p.get("Уке_ФИО", ""))
    return f"{t}-{u}"


def _normalize_slug_from_filename(slug: str) -> str:
    """Суффикс файла после _pos_: tori-uke, с тем же правилом, что и для строки пары."""
    if "-" not in slug:
        return normalize_protocol_token(slug)
    left, right = slug.split("-", 1)
    return f"{normalize_protocol_token(left)}-{normalize_protocol_token(right)}"


def _find_protocol_pair(protocol_file: str, pairs: List[dict]) -> Optional[dict]:
    stem = protocol_file[:-4] if protocol_file.endswith(".csv") else protocol_file
    parsed = CompetitionCSVManager.parse_protocol_filename(stem)
    if parsed:
        _, _, slug_raw = parsed
        ns = _normalize_slug_from_filename(slug_raw)
        for p in pairs:
            k = _pair_slug_from_participants_row(p)
            parts = k.split("-", 1)
            rev = f"{parts[1]}-{parts[0]}" if len(parts) == 2 else k
            if ns == k or ns == rev:
                return p
        return None
    parts = stem.split("_", 2)
    if len(parts) < 3:
        return None
    pair_slug = parts[2]
    ns = _normalize_slug_from_filename(pair_slug)
    for p in pairs:
        k = _pair_slug_from_participants_row(p)
        parts = k.split("-", 1)
        rev = f"{parts[1]}-{parts[0]}" if len(parts) == 2 else k
        if ns == k or ns == rev:
            return p
    return None


def _participant_one_cell(encoded: str) -> str:
    """Одна ячейка: как в протоколе (Имя||детали) или многострочно."""
    d = _decode_participant_cell(encoded)
    if d["detail"]:
        return f"{d['name']}\n{d['detail']}"
    return d["name"]


def _write_tablo_sheet(wb: Workbook, comp_name: str, discipline: str, stage_label: str, final_rows: List[dict], judges: List[dict], pairs: List[dict] = None):
    ws = wb.create_sheet("Результаты")
    judges = judges[:5]
    total_cols = 3 + len(judges) + 2
    end_col_letter = get_column_letter(total_cols)

    # Создаем словарь пар для быстрого доступа
    pairs_by_num = {}
    if pairs:
        for p in pairs:
            try:
                pairs_by_num[int(p.get('номер пары', 0))] = p
            except (ValueError, TypeError):
                pass

    ws.merge_cells(f"A1:{end_col_letter}1")
    ws["A1"] = comp_name
    ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].fill = PatternFill(fill_type="solid", start_color="1A3D5C", end_color="1A3D5C")

    pretty_discipline = discipline.replace("_", "-")
    ws.merge_cells(f"A2:{end_col_letter}2")
    ws["A2"] = pretty_discipline
    ws["A2"].font = Font(size=14, bold=True, color="FFD84A")
    ws["A2"].alignment = Alignment(horizontal="left")
    ws["A2"].fill = PatternFill(fill_type="solid", start_color="224F87", end_color="224F87")

    ws.merge_cells(f"A3:{end_col_letter}3")
    ws["A3"] = stage_label
    ws["A3"].font = Font(size=11, bold=True, color="FFFFFF")
    ws["A3"].alignment = Alignment(horizontal="left")
    ws["A3"].fill = PatternFill(fill_type="solid", start_color="224F87", end_color="224F87")

    # Фон под шапкой, включая область под лого
    for rr in range(1, 4):
        for cc in range(1, total_cols + 1):
            if rr <= 2:
                continue
            ws.cell(rr, cc).fill = PatternFill(fill_type="solid", start_color="224F87", end_color="224F87")

    headers = ["Пара", "Тори", "Уке"] + [f"Судья {j.get('место','')}" for j in judges] + ["Сумма", "Место"]
    ws.append(headers)
    hr = ws.max_row
    hf = PatternFill(fill_type="solid", start_color="2D5A7B", end_color="2D5A7B")
    for i, h in enumerate(headers, 1):
        c = ws.cell(hr, i)
        c.value = h
        c.font = Font(color="FFFFFF", bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = hf

    data_start = ws.max_row + 1
    for row in final_rows:
        judge_vals = []
        for j in judges:
            pos = int(str(j.get("место", 0) or 0))
            v = row.get(f"Судья {pos}", "") if pos else ""
            judge_vals.append(v)

        # Получаем полную информацию об участниках
        pair_num = row.get("номер пары", "")
        tori_cell = row.get("Тори", "")
        uke_cell = row.get("Уке", "")

        # Если нет закодированной строки с ||, кодируем из pair_obj
        if pair_num and pairs_by_num:
            try:
                pair_obj = pairs_by_num.get(int(pair_num))
                if pair_obj:
                    if not tori_cell or "||" not in str(tori_cell):
                        tori_cell = encode_participant_for_protocol(pair_obj, "Тори")
                    if not uke_cell or "||" not in str(uke_cell):
                        uke_cell = encode_participant_for_protocol(pair_obj, "Уке")
            except (ValueError, TypeError):
                pass

        line = [
            pair_num,
            _participant_one_cell(tori_cell),
            _participant_one_cell(uke_cell),
        ] + judge_vals + [row.get("Сумма", ""), row.get("Место", "")]
        ws.append(line)
        r = ws.max_row
        start = 4
        for idx, v in enumerate(judge_vals, start):
            try:
                sv = float(str(v).replace(",", "."))
                ws.cell(r, idx).fill = PatternFill(fill_type="solid", start_color=_theme_gradient(sv), end_color=_theme_gradient(sv))
            except Exception:
                pass
        if str(row.get("Место", "")).strip() in {"1", "2", "3"}:
            ws.cell(r, len(headers)).font = Font(size=16, bold=True, color="CF1F25")

    # логотип (по возможности)
    logo = os.path.join(os.path.dirname(__file__), "static", "federation_logo.png")
    if os.path.exists(logo):
        try:
            from openpyxl.drawing.image import Image  # type: ignore
            img = Image(logo)
            img.width = 120
            img.height = 120
            ws.add_image(img, "H1")
        except Exception:
            pass

    # Границы и авто-таблица
    _apply_grid(ws, 4, ws.max_row, 1, total_cols)

    # Автоматическая высота строк по содержимому
    for r in range(1, ws.max_row + 1):
        max_lines = 1
        for c in range(1, total_cols + 1):
            cell_value = str(ws.cell(r, c).value or '')
            lines = cell_value.count('\n') + 1
            max_lines = max(max_lines, lines)
        # Устанавливаем высоту: базовая 15 + 12 на каждую дополнительную строку
        ws.row_dimensions[r].height = 15 + (max_lines - 1) * 12

    _autosize(ws)

    # Устанавливаем область печати
    ws.print_area = f'A1:{get_column_letter(total_cols)}{ws.max_row}'


def _details_for_protocol_rows(rows: List[dict]) -> List[dict]:
    out = []
    for r in rows:
        dj = str(r.get("details_json", "") or "").strip()
        d = {}
        if dj:
            try:
                d = json.loads(dj)
            except json.JSONDecodeError:
                d = {}
        out.append(d)
    return out


def _write_judge_pair_sheet(
    wb: Workbook,
    sheet_name: str,
    comp_name: str,
    discipline: str,
    stage_label: str,
    judge_name: str,
    pair_name: str,
    techniques: List[str],
    details: List[dict],
):
    ws = wb.create_sheet(sheet_name)
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.merge_cells("A1:K1")
    ws["A1"] = comp_name
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("C3:K3")
    ws["C3"] = f"{discipline}"
    ws["C3"].alignment = Alignment(horizontal="center")
    ws["C3"].font = Font(size=12, bold=True)

    logo = os.path.join(os.path.dirname(__file__), "static", "federation_logo.png")
    if os.path.exists(logo):
        try:
            from openpyxl.drawing.image import Image  # type: ignore
            img = Image(logo)
            img.width = 85
            img.height = 85
            ws.add_image(img, "I1")
        except Exception:
            pass

    ws["A4"] = "Дата:"
    ws["B4"] = datetime.now().strftime("%d.%m.%Y")
    ws["A5"] = "Этап:"
    ws["B5"] = stage_label
    ws["A6"] = "Судья:"
    ws["B6"] = judge_name
    ws["A7"] = "Пара:"
    ws["B7"] = pair_name
    for r in (4, 5, 6, 7):
        ws[f"A{r}"].font = Font(bold=True)

    # Групповой заголовок
    ws.merge_cells("D8:E8")
    ws["D8"] = "Малая ошибка"
    ws["F8"] = "Средняя"
    ws["G8"] = "Большая"
    ws["H8"] = "+/-"
    ws["I8"] = "Забытая"
    ws["J8"] = "SCORE"
    for c in ("D8", "F8", "G8", "H8", "I8", "J8"):
        ws[c].font = Font(bold=True, color="FFFFFF")
        ws[c].alignment = Alignment(horizontal="center", vertical="center")
    ws["D8"].fill = PatternFill(fill_type="solid", start_color="D06262", end_color="D06262")
    ws["E8"].fill = PatternFill(fill_type="solid", start_color="D06262", end_color="D06262")
    ws["F8"].fill = PatternFill(fill_type="solid", start_color="D08B62", end_color="D08B62")
    ws["G8"].fill = PatternFill(fill_type="solid", start_color="CF5A5A", end_color="CF5A5A")
    ws["H8"].fill = PatternFill(fill_type="solid", start_color="5A8BCF", end_color="5A8BCF")
    ws["I8"].fill = PatternFill(fill_type="solid", start_color="9D62CF", end_color="9D62CF")
    ws["J8"].fill = PatternFill(fill_type="solid", start_color="FFFFFF", end_color="FFFFFF")
    ws["J8"].font = Font(bold=True, color="1F2937")

    headers = [
        "№",
        "",
        "TECHNIQUES",
        "Малая 1",
        "Малая 2",
        "Средняя",
        "Большая",
        "+/-",
        "Забытая",
        "Счет",
        "",
    ]
    row_h = 9
    for i, h in enumerate(headers, 1):
        c = ws.cell(row_h, i)
        c.value = h
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(fill_type="solid", start_color="2D5A7B", end_color="2D5A7B")
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Базовые значения
    ws["D10"] = "-1"
    ws["E10"] = "-1"
    ws["F10"] = "-3"
    ws["G10"] = "-5"
    ws["H10"] = "±0.5"
    ws["I10"] = "-10"
    for col in "DEFGHI":
        ws[f"{col}10"].font = Font(italic=True)
        ws[f"{col}10"].alignment = Alignment(horizontal="center")

    m1 = m2 = med = big = pm = fgt = 0
    total_score = 0.0
    start = 11
    tech_fill_a = PatternFill(fill_type="solid", start_color="EAF2FF", end_color="EAF2FF")
    tech_fill_b = PatternFill(fill_type="solid", start_color="F4FAE8", end_color="F4FAE8")
    err_fills = {
        4: PatternFill(fill_type="solid", start_color="FFE8E8", end_color="FFE8E8"),  # m1
        5: PatternFill(fill_type="solid", start_color="FFE8E8", end_color="FFE8E8"),  # m2
        6: PatternFill(fill_type="solid", start_color="FFF4DF", end_color="FFF4DF"),  # med
        7: PatternFill(fill_type="solid", start_color="FFDCDC", end_color="FFDCDC"),  # big
        8: PatternFill(fill_type="solid", start_color="E9F7FF", end_color="E9F7FF"),  # +/-
        9: PatternFill(fill_type="solid", start_color="FDEBFF", end_color="FDEBFF"),  # forgotten
    }
    for idx, tech in enumerate(techniques, 1):
        d = details[idx - 1] if idx - 1 < len(details) else {}
        sc = _score_from_detail(d)
        total_score += sc

        ws.cell(start, 1).value = idx
        ws.cell(start, 3).value = tech
        ws.cell(start, 3).fill = tech_fill_a if idx % 2 else tech_fill_b

        if float(d.get("m1", 0) or 0) > 0:
            ws.cell(start, 4).value = "x"
            m1 += 1
        if float(d.get("m2", 0) or 0) > 0:
            ws.cell(start, 5).value = "x"
            m2 += 1
        if float(d.get("med", 0) or 0) > 0:
            ws.cell(start, 6).value = "x"
            med += 1
        if float(d.get("big", 0) or 0) > 0:
            ws.cell(start, 7).value = "x"
            big += 1

        cp = float(d.get("c_plus", 0) or 0)
        cm = float(d.get("c_minus", 0) or 0)
        if cp != 0 or cm != 0:
            ws.cell(start, 8).value = "+" if cp < 0 else "-"
            pm += 1

        if bool(d.get("forgotten", False)):
            ws.cell(start, 9).value = "x"
            fgt += 1

        ws.cell(start, 10).value = round(sc, 1)
        ws.cell(start, 10).fill = PatternFill(fill_type="solid", start_color="FFFFFF", end_color="FFFFFF")
        for cidx in [1, 4, 5, 6, 7, 8, 9, 10]:
            ws.cell(start, cidx).alignment = Alignment(horizontal="center")
        for cidx, f in err_fills.items():
            ws.cell(start, cidx).fill = f
        start += 1

    ws.cell(start, 3).value = "TOTAL"
    ws.cell(start, 3).font = Font(bold=True)
    ws.cell(start, 4).value = m1
    ws.cell(start, 5).value = m2
    ws.cell(start, 6).value = med
    ws.cell(start, 7).value = big
    ws.cell(start, 8).value = pm
    ws.cell(start, 9).value = fgt
    ws.cell(start, 10).value = round(total_score / 2.0 if fgt > 0 else total_score, 1)
    ws.cell(start, 10).fill = PatternFill(fill_type="solid", start_color="FFE08A", end_color="FFE08A")
    for cidx in [4, 5, 6, 7, 8, 9, 10]:
        ws.cell(start, cidx).font = Font(bold=True)
        ws.cell(start, cidx).alignment = Alignment(horizontal="center")

    # Обводка и вертикальные разделители
    _apply_grid(ws, 8, start, 1, 11)
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 3
    ws.column_dimensions["C"].width = 34
    for col in ["D", "E", "F", "G", "H", "I", "J", "K"]:
        ws.column_dimensions[col].width = 10
    for r in range(8, start + 1):
        ws.row_dimensions[r].height = 22

    _autosize(ws)


def escape_xmlish(s: str) -> str:
    from xml.sax.saxutils import escape

    return escape(str(s or ""))


def _pdf_gradient_background(canvas, width: float, height: float) -> None:
    """Вертикальный «фейд» (несколько стопов сине-голубого, как шапка табло)."""
    n = 36
    stops = [
        (26, 61, 92),
        (34, 79, 135),
        (45, 90, 150),
        (52, 100, 165),
    ]
    for i in range(n):
        t = i / max(1, n - 1)
        seg = t * (len(stops) - 1)
        j = int(seg)
        lt = seg - j
        j = min(j, len(stops) - 2)
        r1, g1, b1 = stops[j]
        r2, g2, b2 = stops[j + 1]
        r = (r1 + (r2 - r1) * lt) / 255.0
        g = (g1 + (g2 - g1) * lt) / 255.0
        b = (b1 + (b2 - b1) * lt) / 255.0
        canvas.setFillColorRGB(r, g, b)
        y0 = height * (1 - (i + 1) / n)
        canvas.rect(0, y0, width, height / n + 1, fill=1, stroke=0)


def _build_pdf(
    file_path: str,
    comp_path: str,
    comp_name: str,
    discipline: str,
    stage_label: str,
    final_rows: List[dict],
    judge_pair_pages: List[Dict[str, Any]],
):
    fn, fnb = _ensure_pdf_fonts()
    styles = getSampleStyleSheet()
    title = ParagraphStyle(
        "t",
        parent=styles["Title"],
        fontName=fnb,
        fontSize=16,
        textColor=colors.white,
        spaceAfter=4,
    )
    sub = ParagraphStyle(
        "sub",
        parent=styles["Normal"],
        fontName=fn,
        fontSize=10,
        textColor=colors.HexColor("#FFD84A"),
    )
    body = ParagraphStyle("b", parent=styles["Normal"], fontName=fn, fontSize=8, leading=10)
    cell_p = ParagraphStyle("cell", parent=styles["Normal"], fontName=fn, fontSize=7, leading=9)

    banner_abs = ""
    try:
        cfg_fp = os.path.join(comp_path, "config.json")
        if os.path.isfile(cfg_fp):
            with open(cfg_fp, "r", encoding="utf-8") as f:
                cfg = json.load(f)
            b = (cfg.get("banner") or "").strip()
            if b:
                cand = b if os.path.isabs(b) else os.path.join(comp_path, b)
                if os.path.isfile(cand):
                    banner_abs = cand
    except Exception:
        pass

    def _cell_html(val: str) -> Paragraph:
        from xml.sax.saxutils import escape

        raw = str(val or "")
        if "||" in raw:
            a, _, b = raw.partition("||")
            inner = f"{escape(a.strip())}<br/>{escape(b.strip())}"
        else:
            inner = escape(raw)
        return Paragraph(inner, cell_p)

    page_size = landscape(A4)
    margin = 12 * mm

    def on_page(canvas, doc):
        w, h = page_size
        canvas.saveState()
        _pdf_gradient_background(canvas, w, h)
        logo = os.path.join(os.path.dirname(__file__), "static", "federation_logo.png")
        if os.path.exists(logo):
            try:
                canvas.drawImage(
                    ImageReader(logo),
                    w - 40 * mm,
                    h - 28 * mm,
                    width=32 * mm,
                    height=24 * mm,
                    preserveAspectRatio=True,
                    mask="auto",
                )
            except Exception:
                pass
        if banner_abs:
            try:
                canvas.drawImage(
                    ImageReader(banner_abs),
                    margin,
                    h - 26 * mm,
                    width=52 * mm,
                    height=20 * mm,
                    preserveAspectRatio=True,
                    mask="auto",
                )
            except Exception:
                pass
        canvas.setStrokeColor(colors.HexColor("#FFFFFF"))
        canvas.setLineWidth(0.8)
        canvas.line(margin, h - 32 * mm, w - margin, h - 32 * mm)
        canvas.restoreState()

    doc = SimpleDocTemplate(
        file_path,
        pagesize=page_size,
        rightMargin=margin,
        leftMargin=margin,
        topMargin=38 * mm,
        bottomMargin=12 * mm,
    )

    pretty_disc = get_discipline_display_name(discipline)
    story = [
        Paragraph(escape_xmlish(comp_name), title),
        Spacer(1, 2 * mm),
        Paragraph(f"{escape_xmlish(pretty_disc)} · {escape_xmlish(stage_label)}", sub),
        Spacer(1, 6 * mm),
    ]

    headers = ["Пара", "Тори", "Уке", "Сумма", "Место"]
    rows_pdf: List[List[Any]] = [headers]
    for r in final_rows:
        t_raw = str(r.get("Тори", "")).strip()
        u_raw = str(r.get("Уке", "")).strip()
        if not t_raw:
            td = _decode_participant_cell(r.get("Тори", ""))
            t_raw = f"{td['name']}||{td['detail']}" if td["detail"] else td["name"]
        if not u_raw:
            ud = _decode_participant_cell(r.get("Уке", ""))
            u_raw = f"{ud['name']}||{ud['detail']}" if ud["detail"] else ud["name"]
        rows_pdf.append(
            [
                str(r.get("номер пары", "")),
                _cell_html(t_raw),
                _cell_html(u_raw),
                str(r.get("Сумма", "")),
                str(r.get("Место", "")),
            ]
        )
    t = Table(rows_pdf, repeatRows=1, colWidths=[18 * mm, 62 * mm, 62 * mm, 22 * mm, 18 * mm])
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2D5A7B")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), fnb),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#94A3B8")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#F8FAFC"), colors.white]),
            ]
        )
    )
    story.append(t)
    story.append(Spacer(1, 8 * mm))

    for p in judge_pair_pages:
        story.append(Paragraph(f"<b>Судья:</b> {escape_xmlish(p['judge'])} &nbsp;|&nbsp; <b>Пара:</b> {escape_xmlish(p['pair'])}", body))
        story.append(Spacer(1, 2 * mm))
        dh = ["№", "Техника", "m1", "m2", "med", "big", "+/-", "forgotten", "score"]
        dr = p["rows"]
        td = Table([dh] + dr, repeatRows=1)
        td.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2D5A7B")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), fnb),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#94A3B8")),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#F1F5F9"), colors.white]),
                ]
            )
        )
        story.append(td)
        story.append(Spacer(1, 5 * mm))

    doc.build(story, onFirstPage=on_page, onLaterPages=on_page)


def _collect_discipline_payload(comp_path: str, comp_display_name: str, dk: str, technique_map: Dict[str, List[str]]) -> Dict[str, Any]:
    cfg = _load_stage(comp_path, dk)
    stage = "final" if cfg.get("current_stage") == "final" else "prelim"
    stage_label = "ФИНАЛ" if stage == "final" else "ПРЕДВАРИТЕЛЬНЫЕ ВСТРЕЧИ"

    sp = _stage_paths(comp_path, dk, stage)
    pairs = _read_csv_rows(sp["pairs"])
    # Читаем финальный протокол ТОЛЬКО из правильного пути для текущего этапа
    final_rows = _read_csv_rows(sp["final"])

    judges = _read_csv_rows(os.path.join(comp_path, dk, "judges_list.csv"))
    judges_sorted = sorted(judges, key=lambda x: int(str(x.get("место", 9999) or 9999)))
    judges_for_tablo = judges_sorted[:5]

    pair_by_slug: Dict[str, dict] = {}
    for p in pairs:
        k = _pair_slug_from_participants_row(p)
        pair_by_slug[k] = p
        pr = k.split("-", 1)
        if len(pr) == 2:
            pair_by_slug[f"{pr[1]}-{pr[0]}"] = p

    judge_pair_pages = []
    # Явно используем путь из stage_paths для текущего этапа
    protocols_dir = sp["protocols"]
    if not os.path.isdir(protocols_dir):
        # Fallback на старую структуру только если новой нет
        protocols_dir = os.path.join(comp_path, dk, "protocols")
        if not os.path.isdir(protocols_dir):
            print(f"⚠️ Протоколы судей не найдены для {dk} ({stage})")
            protocols_dir = None

    if protocols_dir and os.path.isdir(protocols_dir):
        for pf in sorted(os.listdir(protocols_dir)):
            if not pf.endswith(".csv"):
                continue
            stem = pf[:-4]
            parsed = CompetitionCSVManager.parse_protocol_filename(stem)
            if parsed:
                jn, jp, slug_raw = parsed
            else:
                legacy = stem.split("_", 2)
                if len(legacy) < 3:
                    continue
                jn, jp, slug_raw = legacy[0], legacy[1], legacy[2]
            ns = _normalize_slug_from_filename(slug_raw)
            pair = pair_by_slug.get(ns)
            if not pair:
                pair = _find_protocol_pair(pf, pairs)
            if not pair:
                continue
            pair_num = str(pair.get("номер пары", ""))
            pair_name = f"#{pair_num}: {pair.get('Тори_ФИО','')} - {pair.get('Уке_ФИО','')}"
            rows_csv = _read_csv_rows(os.path.join(protocols_dir, pf))
            details = _details_for_protocol_rows(rows_csv)
            techniques = technique_map.get(dk, [r.get("техника", "") for r in rows_csv])

            table_rows = []
            for i, tech in enumerate(techniques, 1):
                d = details[i - 1] if i - 1 < len(details) else {}
                cp = float(d.get("c_plus", 0) or 0)
                cm = float(d.get("c_minus", 0) or 0)
                pm = "+" if cp < 0 else ("-" if cm > 0 else "")
                table_rows.append([
                    i,
                    tech,
                    "x" if float(d.get("m1", 0) or 0) > 0 else "",
                    "x" if float(d.get("m2", 0) or 0) > 0 else "",
                    "x" if float(d.get("med", 0) or 0) > 0 else "",
                    "x" if float(d.get("big", 0) or 0) > 0 else "",
                    pm,
                    "x" if bool(d.get("forgotten", False)) else "",
                    round(_score_from_detail(d), 1),
                ])
            judge_pair_pages.append({
                "judge": f"{jn} (м. {jp})",
                "pair": pair_name,
                "rows": table_rows,
                "techniques": techniques,
                "details": details,
            })

    return {
        "discipline": dk,
        "stage": stage,
        "stage_label": stage_label,
        "comp_name": comp_display_name,
        "pairs": pairs,
        "judges": judges_for_tablo,
        "final_rows": final_rows,
        "judge_pair_pages": judge_pair_pages,
    }


def _save_discipline_protocol(comp_path: str, payload: Dict[str, Any]) -> Dict[str, str]:
    dk = payload["discipline"]
    stage = payload["stage"]
    stage_label = payload["stage_label"]

    # Определяем папку для сохранения протоколов
    if stage == "final":
        # Финальные протоколы сохраняем в подпапку final/
        save_dir = os.path.join(comp_path, dk, "final")
        name_base = f"{dk} protokol ФИНАЛ"
    else:
        # Предварительные протоколы в корне дисциплины
        save_dir = os.path.join(comp_path, dk)
        name_base = f"{dk} protokol"

    os.makedirs(save_dir, exist_ok=True)

    xlsx_path = os.path.join(save_dir, f"{name_base}.xlsx")
    pdf_path = os.path.join(save_dir, f"{name_base}.pdf")

    wb = Workbook()
    wb.remove(wb.active)
    _write_tablo_sheet(wb, payload["comp_name"], dk, stage_label, payload["final_rows"], payload["judges"], payload.get("pairs"))

    used = {"ТАБЛО"}
    for page in payload["judge_pair_pages"]:
        # имя листа: judge+pair
        raw = f"{page['judge']}_{page['pair']}"
        sn = _safe_sheet_name(raw, used)
        _write_judge_pair_sheet(
            wb,
            sn,
            payload["comp_name"],
            dk,
            stage_label,
            page["judge"],
            page["pair"],
            page["techniques"],
            page["details"],
        )

    wb.save(xlsx_path)

    _build_pdf(
        pdf_path,
        comp_path,
        payload["comp_name"],
        dk,
        stage_label,
        payload["final_rows"],
        payload["judge_pair_pages"],
    )
    return {"xlsx": xlsx_path, "pdf": pdf_path}


def protocol_readiness(comp_path: str) -> dict:
    out = []
    for dk in _discipline_folders(comp_path):
        cfg = _load_stage(comp_path, dk)
        stage = "final" if cfg.get("current_stage") == "final" else "prelim"
        sp = _stage_paths(comp_path, dk, stage)
        rows = _read_csv_rows(sp["final"])
        with_sum = sum(1 for r in rows if str(r.get("Сумма", "") or "").strip())
        n_proto = 0
        if os.path.isdir(sp["protocols"]):
            n_proto = len([f for f in os.listdir(sp["protocols"]) if f.endswith(".csv")])
        out.append(
            {
                "key": dk,
                "pairs_registered": len(rows),
                "final_with_total": with_sum,
                "judge_protocol_files": n_proto,
                "ready": with_sum > 0,
            }
        )
    return {"disciplines": out, "overall_ready": any(x["ready"] for x in out) if out else False}


def generate_competition_protocols(
    comp_path: str,
    comp_name: str,
    discipline_key: Optional[str] = None,
    technique_map: Optional[Dict[str, List[str]]] = None,
) -> dict:
    try:
        cfg_file = os.path.join(comp_path, "config.json")
        if not os.path.exists(cfg_file):
            return {"success": False, "message": "Конфигурация соревнования не найдена"}
        with open(cfg_file, "r", encoding="utf-8") as f:
            cfg = json.load(f)
        comp_display_name = cfg.get("name", comp_name)

        tmap = technique_map or {}
        dlist = [discipline_key] if discipline_key else _discipline_folders(comp_path)
        if not dlist:
            return {"success": False, "message": "Нет дисциплин для генерации"}

        generated: Dict[str, Any] = {"disciplines": {}, "results": []}
        for dk in dlist:
            if dk not in _discipline_folders(comp_path):
                continue
            payload = _collect_discipline_payload(comp_path, comp_display_name, dk, tmap)
            files = _save_discipline_protocol(comp_path, payload)
            generated["disciplines"][dk] = [files["xlsx"], files["pdf"], files["csv"]]

        # итоги в results/: копии ТОЛЬКО финальных протоколов каждой дисциплины + индекс
        results_dir = os.path.join(comp_path, "results")
        os.makedirs(results_dir, exist_ok=True)
        copied_files: List[str] = []
        for dk, files in generated["disciplines"].items():
            # Проверяем, что это финальный этап
            cfg = _load_stage(comp_path, dk)
            current_stage = cfg.get('current_stage', 'final')
            # Копируем только если текущий этап - финал (только xlsx и pdf, без csv)
            if current_stage == 'final':
                for fp in files:
                    if not os.path.isfile(fp):
                        continue
                    ext = os.path.splitext(fp)[1].lower()
                    if ext not in (".xlsx", ".pdf"):
                        continue
                    dst = os.path.join(results_dir, os.path.basename(fp))
                    shutil.copy2(fp, dst)
                    copied_files.append(dst)

        index_xlsx = os.path.join(results_dir, "results_index.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append(["Дисциплина", "Файлы"]) 
        for dk, files in generated["disciplines"].items():
            ws.append([dk, " ; ".join(files)])
        wb.save(index_xlsx)
        generated["results"] = [index_xlsx] + copied_files

        return {"success": True, "generated": generated}
    except Exception as e:
        return {"success": False, "message": f"Ошибка при генерации протоколов: {e}"}
